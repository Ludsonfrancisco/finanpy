import openpyxl
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from accounts.models import Account
from categories.models import Category
from transactions.models import Transaction
from datetime import datetime

User = get_user_model()

class Command(BaseCommand):
    help = 'Importa dados financeiros de um arquivo Excel (.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, help='Caminho para o arquivo .xlsx')
        parser.add_argument('--user-email', type=str, help='Email do usuário para quem os dados serão importados')

    def handle(self, *args, **options):
        file_path = options['file']
        user_email = options['user_email']

        if not file_path or not user_email:
            self.stderr.write('É necessário fornecer --file e --user-email')
            return

        try:
            user = User.objects.get(email=user_email)
        except User.DoesNotExist:
            self.stderr.write(f'Usuário com email {user_email} não encontrado.')
            return

        wb = openpyxl.load_workbook(file_path)

        # 1. Importar Contas
        self.stdout.write('Importando contas...')
        ws_accounts = wb['Contas']
        accounts_map = {}
        for row in ws_accounts.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            name, acc_type, balance, currency = row
            account, created = Account.objects.get_or_create(
                user=user,
                name=name,
                defaults={
                    'type': acc_type,
                    'initial_balance': Decimal(str(balance or 0)),
                    'currency': currency or 'BRL'
                }
            )
            accounts_map[name] = account
            if created:
                self.stdout.write(f'  Conta criada: {name}')

        # 2. Importar Categorias
        self.stdout.write('Importando categorias...')
        ws_categories = wb['Categorias']
        categories_map = {}
        for row in ws_categories.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            name, cat_type, color, icon = row
            category, created = Category.objects.get_or_create(
                user=user,
                name=name,
                type=cat_type,
                defaults={
                    'color': color or '#10b981',
                    'icon': icon
                }
            )
            categories_map[name] = category
            if created:
                self.stdout.write(f'  Categoria criada: {name} ({cat_type})')

        # 3. Importar Transações
        self.stdout.write('Importando transações...')
        ws_transactions = wb['Transações']
        count = 0
        for row in ws_transactions.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            date_val, description, amount, acc_name, cat_name, trans_type = row
            
            if isinstance(date_val, str):
                date_obj = datetime.strptime(date_val, '%Y-%m-%d').date()
            else:
                date_obj = date_val.date() if isinstance(date_val, datetime) else date_val

            account = accounts_map.get(acc_name)
            category = categories_map.get(cat_name)

            if not account or not category:
                self.stderr.write(f'  Erro na linha: Conta "{acc_name}" ou Categoria "{cat_name}" não encontrada.')
                continue

            Transaction.objects.create(
                user=user,
                account=account,
                category=category,
                description=description,
                amount=Decimal(str(amount)),
                date=date_obj,
                type=trans_type
            )
            count += 1
        
        self.stdout.write(self.style.SUCCESS(f'Sucesso! {count} transações importadas.'))
